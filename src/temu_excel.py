"""
Temu 运营助手 — Excel 导出工具（适配 Temu 各模块）
"""
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.run([sys.executable, "-m", "pip", "install", "-q", "openpyxl"], check=True)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="2F6DB5", end_color="2F6DB5", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="微软雅黑", size=10)
ALT_FILL   = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")


def _write_sheet(ws, headers: list, rows: list, sheet_title: str = "数据"):
    ws.title = sheet_title

    # 写表头
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 写数据行
    for row_idx, row in enumerate(rows, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    # 列宽自适应
    for col_idx, _ in enumerate(headers, 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_idx, row in enumerate(rows, 2):
            if col_idx - 1 < len(row):
                val_len = min(len(str(row[col_idx - 1] or '')), 60)
                max_len = max(max_len, val_len)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    # 冻结首行
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def write_temu_excel(output_path: str, sheets: list[dict]):
    """
    写多 sheet Excel
    sheets: [{"title": str, "headers": [...], "rows": [[...], ...]}, ...]
    """
    wb = openpyxl.Workbook()
    # 移除默认 sheet
    wb.remove(wb.active)

    for sheet_def in sheets:
        ws = wb.create_sheet()
        _write_sheet(
            ws,
            headers=sheet_def.get("headers", []),
            rows=sheet_def.get("rows", []),
            sheet_title=sheet_def.get("title", "Sheet")
        )

    # 确保目录存在
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    wb.save(output_path)
    return output_path
