"""
Excel 导出工具
统一处理价格列表 Excel 的生成逻辑，包含兜底补全标记高亮。
"""
from datetime import datetime
from pathlib import Path
from typing import List, Dict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill


# 颜色常量
COLOR_HEADER_BG   = "1F4E79"    # 深蓝表头
COLOR_ROW_ALT     = "EBF3FB"    # 交替行浅蓝
COLOR_FALLBACK_BG = "FFF2CC"    # 兜底补全行 —— 淡黄高亮
COLOR_FALLBACK_FG = "7F6000"    # 兜底补全行字体色（深金）
COLOR_MISSING_BG  = "FFE0E0"    # 仍缺价格行 —— 淡红
COLOR_MISSING_FG  = "9C0006"    # 仍缺价格行字体色（深红）


def write_price_excel(
    sku_list: List[Dict],
    out_dir: Path,
    filename_prefix: str = "price_list",
) -> str:
    """
    将 SKU 列表写成 Excel 并返回文件路径。

    列：款号 / 商品名称 / 页面价 / 原价 / 价格来源 / 商品链接

    颜色规则：
    - 兜底（detail_page）补全的行：淡黄背景
    - 仍无价格的行：淡红背景
    - 普通交替行：白/浅蓝
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file = out_dir / f"{filename_prefix}_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "价格列表"

    headers = ["款号(SKU ID)", "商品名称", "页面价(元)", "原价(元)", "价格来源", "商品链接"]
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, item in enumerate(sku_list, 2):
        source      = item.get("price_source", "")
        has_price   = item.get("current_price") is not None
        is_fallback = source == "detail_page"
        is_missing  = not has_price

        # 行填充色 & 字体
        if is_missing:
            row_fill = PatternFill("solid", fgColor=COLOR_MISSING_BG)
            row_font = Font(color=COLOR_MISSING_FG)
        elif is_fallback:
            row_fill = PatternFill("solid", fgColor=COLOR_FALLBACK_BG)
            row_font = Font(color=COLOR_FALLBACK_FG)
        elif row_idx % 2 == 0:
            row_fill = PatternFill("solid", fgColor=COLOR_ROW_ALT)
            row_font = Font()
        else:
            row_fill = PatternFill()
            row_font = Font()

        # 价格来源可读文本
        source_label = {
            "list_page":   "列表页",
            "detail_page": "详情页(兜底)",
            "":            "未知",
        }.get(source, source)

        cells_values = [
            item.get("sku_id", ""),
            item.get("name", ""),
            item.get("current_price") if has_price else "—",
            item.get("original_price") or "—",
            source_label,
            item.get("product_url", ""),
        ]
        for col, val in enumerate(cells_values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.font = row_font

    # 列宽
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 55
    ws.freeze_panes = "A2"

    # 统计摘要（在最后追加）
    total       = len(sku_list)
    with_price  = sum(1 for r in sku_list if r.get("current_price") is not None)
    fallback_n  = sum(1 for r in sku_list if r.get("price_source") == "detail_page")
    missing_n   = total - with_price
    summary_row = total + 3
    ws.cell(row=summary_row, column=1, value="统计摘要").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value=f"总计：{total} 个 SKU")
    ws.cell(row=summary_row + 2, column=1, value=f"有价格：{with_price} 个")
    ws.cell(row=summary_row + 3, column=1, value=f"其中详情页兜底补全：{fallback_n} 个")
    ws.cell(row=summary_row + 4, column=1, value=f"仍缺价格：{missing_n} 个")

    wb.save(str(out_file))
    return str(out_file)
